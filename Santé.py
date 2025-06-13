import streamlit as st
import pandas as pd
import plotly.express as px
import numpy as np
from streamlit_extras.metric_cards import style_metric_cards # beautify metric card with css
import plotly.graph_objects as go
import base64
import json
import openpyxl
from openpyxl import load_workbook
import altair as alt
from pandas.api.types import (
    is_categorical_dtype,
    is_datetime64_any_dtype,
    is_numeric_dtype,
    is_object_dtype,
)

st. set_page_config(layout="wide")

def read_excel_file(file):
    data = load_workbook(file)
    datas = data.active
    donnees = []
    for ligne in datas.iter_rows(values_only=True):
        donnees.append(list(ligne))
    en_tetes = donnees[0]
    donnees = donnees[1:]
    new_df = pd.DataFrame(donnees, columns=en_tetes)
    return new_df
df=read_excel_file("BD Santé act 13 juin 2025.xlsx")



# Fonction pour lire une image locale et la convertir en base64
def get_base64_image(image_path):
    with open(image_path, "rb") as img_file:
        encoded = base64.b64encode(img_file.read()).decode()
    return encoded

# Spécifiez le chemin de votre image locale
image_path = "images/background.jpeg"  # Remplacez par le chemin de votre image
base64_image = get_base64_image(image_path)

# CSS pour définir l'arrière-plan
page_bg_img = f"""
<style>
[data-testid="stAppViewContainer"] {{
    background-image: url("data:image/jpeg;base64,{base64_image}");
    background-size: cover;
    background-position: center;
    background-repeat: no-repeat;
    background-attachment: no-fixed;
    height: 100vh;
    margin: 0;
    display: flex;

    
    }}
    [data-testid="stSidebar"] {{
        background-color: #000 !important;  /* Fond noir */
        border: 2px solid #f7a900 !important;  /* Bordure rouge */
        border-radius: 10px;  /* Coins arrondis */
        margin-top: 0 px;  /* Ajuster la position vers le haut */
        position: relative;
        z-index: 1;  /* S'assurer que la barre latérale est au-dessus du contenu */
        padding: 10px;
    }}

    [data-testid="stHeader"] {{
    background: rgba(0, 0, 0, 0);
    color: white;
    }}

    [data-testid="stToolbar"] {{
    right: 2rem;
    }}
    </style>
    """

st.markdown(page_bg_img, unsafe_allow_html=True)
st.markdown('<div style="text-align:center;width:100%;"><h1 style="color:white;background-color:black;border:red;border-style:solid;border-radius:5px;">TABLEAU DE BORD DE LA DEMANDE SUR LE MARCHE IVOIRIEN DE LA SANTE </h1></div>', unsafe_allow_html=True)
st.write("")


st.header("Base de données personnalisée",divider="rainbow" )
def filter_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """
    Adds a UI on top of a dataframe to let viewers filter columns

    Args:
        df (pd.DataFrame): Original dataframe

    Returns:
        pd.DataFrame: Filtered dataframe
    """
    modify = st.checkbox("AJOUTEZ UN FILTRE")
    

    if not modify:
        return df

    df = df.copy()

    # Try to convert datetimes into a standard format (datetime, no timezone)
    for col in df.columns:
        if is_object_dtype(df[col]):
            try:
                df[col] = pd.to_datetime(df[col])
            except Exception:
                pass

        if is_datetime64_any_dtype(df[col]):
            df[col] = df[col].dt.tz_localize(None)

    modification_container = st.container()

    with modification_container:
        to_filter_columns = st.multiselect("Choisissez les variables que vous souhaitez utiliser comme filtre", df.columns)
        for column in to_filter_columns:
            left, right = st.columns((1, 20))
            # Treat columns with < 10 unique values as categorical
            int_columns = df.select_dtypes(include="int").columns
            float_columns = df.select_dtypes(include="float").columns

            if is_numeric_dtype(df[column]) :
                _min = int(df[column].min())
                _max = int(df[column].max())
                user_num_input = right.slider(
                    f"Valeurs de {column}",
                    min_value=_min,
                    max_value=_max,
                    value=(_min, _max),
                )
                df = df[df[column].between(*user_num_input)]
            elif is_datetime64_any_dtype(df[column]):
                user_date_input = right.date_input(
                    f"Valeur de {column}",
                    value=(
                        df[column].min(),
                        df[column].max(),
                    ),
                )
                if len(user_date_input) == 2:
                    user_date_input = tuple(map(pd.to_datetime, user_date_input))
                    start_date, end_date = user_date_input
                    df = df.loc[df[column].between(start_date, end_date)]
            elif is_categorical_dtype(df[column]) or df[column].unique().shape[0]<100:
                arr=df[column].unique()
                user_cat_input = right.multiselect(
                    f"Valueur de {column}",
                    arr
                    ,
                    default=list(arr),
                )
                df = df[df[column].isin(user_cat_input)]
            else:
                user_text_input = right.text_input(
                    f"Substring or regex in {column}",
                )
                if user_text_input:
                    df = df[df[column].astype(str).str.contains(user_text_input)]

    return df
    
df_perso=filter_dataframe(df)
st.dataframe(df_perso)

colors = px.colors.sequential.Rainbow_r
colors.extend(px.colors.sequential.Agsunset)
colors.extend(px.colors.sequential.Aggrnyl)


# SECTION GRAPHIQUE
st.header("Analyses graphiques", divider="rainbow")
st.subheader("HISTOGRAMME")

def barmode_selected(t):
    if t =='empilé':
        a='relative'  
    else: 
        a='group'
    return a
st.subheader("ANALYSE CROISEE ENTRE VARIABLES CATEGORIELLES")
selected_variable_1 = st.selectbox('***Variable en abscisse***', ['Catégories des RHS prioritaires', "Sous-catégories"], index=1)
selected_variable_2 = st.selectbox("***Variable en ordonnée***", df.columns, index=3)

# Filtrer les données
df2 = df.dropna(subset=[selected_variable_2])
df2 = df2[df2[selected_variable_2] != 0]

# Convertir la colonne `selected_variable_2` en numérique si possible
df2[selected_variable_2] = pd.to_numeric(df2[selected_variable_2], errors='coerce')
df2 = df2.dropna(subset=[selected_variable_2])  # Supprimer les lignes avec des valeurs non numériques après conversion

if selected_variable_1 == "Sous-catégories":
    selected_variable = st.multiselect(
        "***Sélectionnez la (ou les) sous-catégorie(s)***",
        df2["Sous-catégories"].unique(),
        "IDE"
    )
    df2 = df2[df2["Sous-catégories"].isin(selected_variable)]
    df_sorted = df2.groupby("Catégories des RHS prioritaires")[selected_variable_2].sum().sort_values(ascending=False).reset_index()

    # Assurez-vous que les catégories sont ordonnées correctement
    category_order = df_sorted["Catégories des RHS prioritaires"].tolist()
    fig_croisé = px.histogram(
        df_sorted,
        x='Catégories des RHS prioritaires',
        y=selected_variable_2,
        color='Catégories des RHS prioritaires',
        color_discrete_sequence=colors
    )
    fig_croisé.update_xaxes(categoryorder='array', categoryarray=category_order)
else:
    df_sorted = df2.groupby('Catégories des RHS prioritaires')[selected_variable_2].sum().reset_index()
    fig_croisé = px.histogram(
        df_sorted,
        x='Catégories des RHS prioritaires',
        y=selected_variable_2,
        color='Catégories des RHS prioritaires',
        color_discrete_sequence=colors
    )

fig_croisé.update_layout(title=f'Graphique en barres groupées - {selected_variable_1 } vs {selected_variable_2 }')
fig_croisé.update_layout({'plot_bgcolor': 'rgba(0, 0, 0, 0)','paper_bgcolor': 'rgba(0, 0, 0, 0.3)',},title_x=0.20)
# Mise en forme
fig_croisé.update_layout(
    yaxis_title='Effectif',
    xaxis_tickangle=45,
    legend=dict(
        orientation="h",
        yanchor="top",
        y=-0.2,        # Ajuste la valeur négative si besoin pour déplacer la légende plus bas
        xanchor="center",
        x=0.5
    ),
    margin=dict(b=140),  # Augmente la marge basse pour la légende exportée
    plot_bgcolor='white',
    paper_bgcolor='white'
)
fig_croisé.update_traces(marker=dict(opacity=0.7))
fig_croisé.update_xaxes(showticklabels=False)  # Supprimer les libellés sous les bandes
fig_croisé.update_yaxes(tickformat=".0f")  # Format entier sans 'k' ou 'M'
fig_croisé.update_traces(
    texttemplate='%{y}',  # Affiche les valeurs réelles sans format abrégé
    textposition='outside'  # Place les valeurs au-dessus des barres
)
st.plotly_chart(fig_croisé,use_container_width=True)

quant,cam=st.columns(2,gap='medium')


with quant:
    st.subheader("ANALYSE CROISEE ENTRE VARIABLES NUMERIQUES")
    int_columns = df.select_dtypes(include="int").columns
    float_columns = df.select_dtypes(include="float").columns
    selected_variable_3 = st.selectbox("***Variable en abscisse***", int_columns.union(float_columns))
    selected_variable_4 = st.selectbox("***Variable en ordonnée***",int_columns.union(float_columns),index=2)
    fig_scatter_matrix = px.scatter(df, x=selected_variable_3, y=selected_variable_4)
    fig_scatter_matrix.update_layout(title=f'Nuage de points entre {selected_variable_3} et {selected_variable_4}')
    fig_scatter_matrix.update_layout({'plot_bgcolor': 'rgba(0, 0, 0, 0)','paper_bgcolor': 'rgba(0, 0, 0, 0.3)',},title_x=0.15)
    st.plotly_chart(fig_scatter_matrix, use_container_width=True)

with cam:
    st.subheader("CAMEMBERT")
    selected_categorical_variable_p = st.selectbox("***Sélectionnez la variable***", df.columns[1:], index=0)
    if selected_categorical_variable_p=="Sous-catégories":
        category_counts = df[selected_categorical_variable_p].value_counts()
        fig_pie = px.pie(names=category_counts.index, values=category_counts.values, title=f"Répartition de la variable {selected_categorical_variable_p}",color_discrete_sequence=colors)
    else:
        df3 = df.dropna(subset=[selected_categorical_variable_p])
        df3= df3[df3[selected_categorical_variable_p] != 0]
        selected_variable5=st.multiselect("***Sélectionnez la (ou les) sous-catégorie(s)***", df3["Sous-catégories"].unique())
        df3=df3[df3["Sous-catégories"].isin(selected_variable5)]
        fig_pie = px.pie(names=df3["Catégories des RHS prioritaires"], values=(df3[selected_categorical_variable_p] / df3[selected_categorical_variable_p].sum()) * 100, title=f"Répartition de la variable {selected_categorical_variable_p}",color_discrete_sequence=colors)


    fig_pie.update_layout({'plot_bgcolor': 'rgba(0, 0, 0, 0)','paper_bgcolor': 'rgba(0, 0, 0, 0.3)',},title_x=0.25)
    st.plotly_chart(fig_pie, use_container_width=True)
